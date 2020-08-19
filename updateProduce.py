#! /usr/bin/python3
# updateProduce.py - Correct cost in produce sales spreadsheet

import openpyxl

wb = openpyxl.load_workbook('produce sales.xlsx')
sheet = wb['Hoja1']

# The produce types and their updated prices
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon:': 1.27}

# Loop through the rows and update the price
for rowNum in range(2, sheet.get_highest_row()):  # Skip the first row
    produceName = sheet.cell(row=rowNum, colum=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]
        wb.save('updatedProduceSales.xlsx')
