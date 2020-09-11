#! python3

from tkinter import filedialog as FileDialog
from tkinter import messagebox as MessageBox
from tkinter import *  # Prueba borrar cuadro dialogo
from tkinter.ttk import * # Prueba borrar cuadro dialogo
from os.path import basename
import pyautogui
import pyperclip
import openpyxl
import xlrd
import xlwt
from xlutils.copy import copy
import re
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os
import time

path = r"S:\OPS\Dispatchers\Preestibas"
file = 'PREESTIBAS FISICAS.xlsx'
#style = xlwt.easyxf('pattern: pattern solid, fore_colour yellow') # background color
root = Tk() # Prueba borrar cuadro dialogo
root.update()


def clear():  # ERASE SCREEN
    if os.name == "nt":      # Windows
        os.system("cls")
    else:
        os.system("clear")   # Linux
        

def getFile(path):  # DIALOG TO CHOOSE A FILE
    file = FileDialog.askopenfilename(
        initialdir=path,
        filetypes=(('xlsx', '*.xlsx'), ('all files', '*.*')),
        title='Escoja el archivo de Preestibas a procesar'
     )
    #print('\nEl nombre del archivo inicial es: ' + basename(file) + '\n')
    print (file)
    return(file)


def getPreestibas(file):  # Get a list of Prrestibas
    book = openpyxl.load_workbook(file)
    sheet = book.worksheets[1]  
    preestibas = []
    for cell in sheet['A']:  # Column A
        preestibas.append(cell.value)
    print('Total preestibas digitadas:' + str(len(preestibas)))       
    return(preestibas)



def main():  # MAIN
    #clear()
    file = getFile(path)
    root.destroy() # Prueba de cerrar caja de dialogo
    preestibas = getPreestibas(file)
    #print('Total preestibas digitadas:' + str(len(preestibas)))   

    book = xlrd.open_workbook(file)
    sheet = book.sheet_by_index(1)
    book = copy(book)
    sheet2 = book.get_sheet(1)  # same sheet in order to can save

    for i in range(sheet.nrows):  # Loop through the sheet
        contenedor = (repr(sheet.cell_value(i, 0)).replace("'", ""))
        if type(preestibas) == list:
            if contenedor in preestibas:
                #pyperclip.copy(contenedor)
                pyautogui.click(702, 384) # clic contenedor Monitor 1
                pyautogui.typewrite(contenedor[:-1])
                pyautogui.click(2622, 375) # clic contenedor Monitor 2
                pyautogui.typewrite(contenedor[:-1])
                #time.sleep(1)
                pyautogui.click(1081, 606) # clic aceptar Monitor 1
                pyautogui.click(3004, 606) # clic aceptar Monitor 2
                print(contenedor)
                #sheet2.write(i, 2, contenedor, style)  # Modify cell
                

    pyautogui.click(1168, 606) # clic Cancelar Monitor 1
    pyautogui.click(3095, 606) # clic Cancelar Monitor 2
    

if __name__ == '__main__':
    main()


os.system("pause") # Press a key to continue 
