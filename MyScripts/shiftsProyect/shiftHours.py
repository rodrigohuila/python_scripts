#! /usr/bin/python3
import datetime
import openpyxl

#date = input('Ingrese la fecha: ')
#date = (2020, 5, 20)

def run():
    turnos_horarios = {
        'M': 'Mañana de 06:00 a 14:00',
        'T': 'Tarde de 14:00 a 22:00',
        'N': 'Noche de 22:00 a 06:00',
        'M12': 'Mañana12h de 06:00 a 18:00',
        'N12': 'Noche12h de 22:00 a 06:00',
        'K': '10hT de 06:00 a 16:00',
        'LIB': 'Libre',
        'VAC' : 'Vacaciones'
    }

    for turno in turnos_horarios.items():
        print(turno)
    
    #print(turnos_horarios)

    #for turno in turnos_horarios.keys():
    #     print(turno)

#if __name__ == '__main__':
#    run()

# Open WorkBook Excel
wb = openpyxl.load_workbook('Turnos.xlsx')
sheet = wb['Cuadrante']
c = sheet['A3']
#print (c.value)
#print ('Row ' + str(c.row) + ', Column ' + str(c.column) + ' is ' + c.value)
#print ('Cell ' + c.coordinate + ' is ' + c.value)


shifts = []

#Range (n hasta n, de n en n)
for i in range(2, 33, 1):
    #print(i,sheet.cell(row=3, column=i).value)
    shifts.append(sheet.cell(row=3, column=i).value)

print (len(shifts))

#for i in range(0, len(shifts)):
  #print (shifts[0]) 

## Modificar libro
 
# Create Shift in the last position
wb.create_sheet() 
# Create Shift in the psosition we want 
wb.create_sheet(index=0, title='First Sheet')
# Remove shift
del wb['First Sheet']
print (wb.sheetnames)

wb.create_sheet(index=0, title='Horarios')
sheet2 = wb['Horarios']
sheet2['A1'] = 'Hello world!'
sheet2['A1'].value